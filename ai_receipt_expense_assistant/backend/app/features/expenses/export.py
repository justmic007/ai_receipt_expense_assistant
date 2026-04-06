import io
from datetime import datetime
from sqlalchemy.orm import Session
from app.features.expenses.models import Expense
from app.features.receipts.models import Receipt
import uuid


def get_expenses_in_range(
    db: Session, user_id: str, from_date: str, to_date: str
) -> list:
    query = db.query(Expense).filter(Expense.user_id == uuid.UUID(user_id))
    if from_date:
        query = query.filter(Expense.expense_date >= from_date)
    if to_date:
        query = query.filter(Expense.expense_date <= to_date)
    return query.order_by(Expense.expense_date.desc()).all()


def get_receipts_in_range(
    db: Session, user_id: str, from_date: str, to_date: str
) -> list:
    query = db.query(Receipt).filter(Receipt.user_id == uuid.UUID(user_id))
    if from_date:
        query = query.filter(Receipt.receipt_date >= from_date)
    if to_date:
        query = query.filter(Receipt.receipt_date <= to_date)
    return query.order_by(Receipt.receipt_date.desc()).all()


def export_expenses_excel(
    db: Session, user_id: str, from_date: str, to_date: str
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    expenses = get_expenses_in_range(db, user_id, from_date, to_date)
    receipts = get_receipts_in_range(db, user_id, from_date, to_date)

    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="0F172A")
    accent_fill = PatternFill("solid", fgColor="2563EB")
    light_fill = PatternFill("solid", fgColor="F8FAFC")

    ws_summary["A1"] = "ReceiptAI — Expense Report"
    ws_summary["A1"].font = Font(bold=True, size=14, color="0F172A")
    ws_summary["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}"
    ws_summary["A2"].font = Font(size=10, color="94A3B8")
    ws_summary["A3"] = f"Period: {from_date or 'All time'} to {to_date or 'Today'}"
    ws_summary["A3"].font = Font(size=10, color="94A3B8")

    ws_summary.append([])

    total_spent = sum(e.amount for e in expenses)
    stats = [
        ["Metric", "Value"],
        ["Total Expenses", len(expenses)],
        [
            "Total Spent",
            f"{expenses[0].currency if expenses else 'NGN'} {total_spent:,.2f}",
        ],
        ["Total Receipts", len(receipts)],
        ["Date Range", f"{from_date or 'All'} → {to_date or 'Today'}"],
    ]

    for i, row in enumerate(stats):
        ws_summary.append(row)
        if i == 0:
            for col in range(1, 3):
                cell = ws_summary.cell(row=ws_summary.max_row, column=col)
                cell.font = header_font
                cell.fill = header_fill

    ws_summary.append([])
    ws_summary.append(["Category Breakdown", "", ""])

    cat_header_row = ws_summary.max_row
    for col, label in enumerate(["Category", "Count", "Total Amount"], 1):
        cell = ws_summary.cell(row=cat_header_row, column=col)
        cell.value = label
        cell.font = header_font
        cell.fill = accent_fill

    category_map = {}
    for e in expenses:
        cat = e.category or "Other"
        if cat not in category_map:
            category_map[cat] = {"count": 0, "total": 0}
        category_map[cat]["count"] += 1
        category_map[cat]["total"] += e.amount

    for cat, vals in sorted(category_map.items(), key=lambda x: -x[1]["total"]):
        currency = expenses[0].currency if expenses else "NGN"
        ws_summary.append([cat, vals["count"], f"{currency} {vals['total']:,.2f}"])

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 20
    ws_summary.column_dimensions["C"].width = 25

    # ── Sheet 2: Expenses ─────────────────────────────────────────
    ws_expenses = wb.create_sheet("Expenses")
    exp_headers = [
        "Date",
        "Merchant",
        "Category",
        "Amount",
        "Currency",
        "Note",
        "Reviewed",
    ]

    for col, header in enumerate(exp_headers, 1):
        cell = ws_expenses.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for expense in expenses:
        ws_expenses.append(
            [
                expense.expense_date or "",
                expense.merchant_name or "",
                expense.category or "",
                expense.amount,
                expense.currency or "NGN",
                expense.note or "",
                "Yes" if expense.is_reviewed else "No",
            ]
        )

    col_widths = [15, 35, 20, 15, 10, 30, 12]
    for i, width in enumerate(col_widths, 1):
        ws_expenses.column_dimensions[get_column_letter(i)].width = width

    # ── Sheet 3: Receipts ─────────────────────────────────────────
    ws_receipts = wb.create_sheet("Receipts")
    rec_headers = [
        "Date",
        "Merchant",
        "Total Amount",
        "Currency",
        "Category",
        "Status",
        "AI Model",
        "Filename",
    ]

    for col, header in enumerate(rec_headers, 1):
        cell = ws_receipts.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for receipt in receipts:
        ws_receipts.append(
            [
                receipt.receipt_date or "",
                receipt.merchant_name or "",
                receipt.total_amount or 0,
                receipt.currency or "NGN",
                receipt.category or "",
                receipt.status,
                receipt.model_used or "",
                receipt.original_filename,
            ]
        )

    rec_widths = [15, 35, 15, 10, 20, 12, 30, 40]
    for i, width in enumerate(rec_widths, 1):
        ws_receipts.column_dimensions[get_column_letter(i)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def export_expenses_pdf(
    db: Session, user_id: str, from_date: str, to_date: str
) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
    )

    expenses = get_expenses_in_range(db, user_id, from_date, to_date)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm
    )

    styles = getSampleStyleSheet()
    navy = colors.HexColor("#0F172A")
    blue = colors.HexColor("#2563EB")
    light = colors.HexColor("#F8FAFC")
    muted = colors.HexColor("#94A3B8")

    title_style = ParagraphStyle(
        "title", fontSize=20, fontName="Helvetica-Bold", textColor=navy, spaceAfter=4
    )
    sub_style = ParagraphStyle(
        "sub", fontSize=10, fontName="Helvetica", textColor=muted, spaceAfter=2
    )
    section_style = ParagraphStyle(
        "section",
        fontSize=13,
        fontName="Helvetica-Bold",
        textColor=navy,
        spaceBefore=12,
        spaceAfter=6,
    )

    story = []

    story.append(Paragraph("ReceiptAI", title_style))
    story.append(Paragraph("Expense Report", sub_style))
    story.append(
        Paragraph(
            f"Period: {from_date or 'All time'} → {to_date or 'Today'}", sub_style
        )
    )
    story.append(
        Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}", sub_style)
    )
    story.append(Spacer(1, 10 * mm))

    total_spent = sum(e.amount for e in expenses)
    currency = expenses[0].currency if expenses else "NGN"

    summary_data = [
        ["Total Expenses", "Total Spent", "Date Range"],
        [
            str(len(expenses)),
            f"{currency} {total_spent:,.2f}",
            f"{from_date or 'All'} → {to_date or 'Today'}",
        ],
    ]

    summary_table = Table(summary_data, colWidths=[55 * mm, 75 * mm, 55 * mm])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), navy),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BACKGROUND", (0, 1), (-1, 1), light),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [light]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(summary_table)
    story.append(Spacer(1, 8 * mm))

    story.append(Paragraph("Expense Details", section_style))

    if not expenses:
        story.append(Paragraph("No expenses found for this date range.", sub_style))
    else:
        exp_data = [["Date", "Merchant", "Category", "Amount", "Note"]]
        for e in expenses:
            exp_data.append(
                [
                    e.expense_date or "—",
                    (e.merchant_name or "—")[:30],
                    e.category or "—",
                    f"{e.currency} {e.amount:,.2f}",
                    (e.note or "—")[:20],
                ]
            )

        exp_data.append(["", "TOTAL", "", f"{currency} {total_spent:,.2f}", ""])

        exp_table = Table(
            exp_data, colWidths=[25 * mm, 55 * mm, 35 * mm, 35 * mm, 35 * mm]
        )
        exp_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), blue),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("BACKGROUND", (0, -1), (-1, -1), light),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, light]),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                    ("ALIGN", (3, 0), (3, -1), "RIGHT"),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story.append(exp_table)

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()
